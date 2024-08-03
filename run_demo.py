import os
import argparse
import math
import openpyxl
import json

parser = argparse.ArgumentParser()
parser.add_argument("-p", default="traces", help="trace path")
parser.add_argument("-n", default="Sheet", help="sheet name")

def run(path, name):

    layer_dict = {"createQKV":'B', "QK":'C', "SV":'D', "Wo":'E', "L1":'F', "L2":'G'}
    if os.path.exists('result.xlsx'):
        wb = openpyxl.load_workbook('result.xlsx')
    else:
        wb = openpyxl.Workbook()

    fin = open("models_s", 'r')
    lines = fin.readlines()
    for line in lines:
        model, n_layers, d_model, n_heads, d_head, par, PP = line.strip().split(' ')
        par = int(par)
        '''
        sum_path = "traces/" + model + "/SUM"
        slist = sorted(os.listdir(sum_path))
        name = model + "_SUM"
        if name in wb.get_sheet_names():
            wb.remove_sheet(wb.get_sheet_by_name(name))

        wb.create_sheet(title=name)
        sh = wb.get_sheet_by_name(name)

        sh['A'+str(1)] = 'batch_multicolumns'
        sh['B'+str(1)] = 'createQKV'
        sh['C'+str(1)] = 'QK'
        sh['D'+str(1)] = 'SV'
        sh['E'+str(1)] = 'Wo'
        sh['F'+str(1)] = 'L1'
        sh['G'+str(1)] = 'L2'
        sh['H'+str(1)] = 'total cycles'
        for i, s in enumerate(slist):
            sh['A'+str(i+2)] = s
            # bs = s.split('_')[0]

            flist = sorted(os.listdir(sum_path + '/' + s))
            for j, f in enumerate(flist):
                print(model, s, f)
                if f == "QK" or f == "SV":
                    continue

                os.system("echo " + f + " >> logs/" + model + "_SUM_" + s + ".log")
                os.system("./build/dramsim3main configs/HBM2_8Gb_x128.ini -c 10000000 -t " + sum_path + '/' + s + '/' + f  + " >> logs/" + model + "_SUM_" + s + ".log")

                with open('dramsim3.json') as df:
                    json_object = json.load(df)
                    cycles = json_object['0']['num_cycles']
                    if cycles == 10000000:
                        raise Exception("compute not finished!")
                    cycles *= int(n_layers)
                    if f == "createQKV":
                        cycles *= 3
                    elif f == "QK" or f == "SV":
                        cycles *= int(n_heads)/par
                    sh[layer_dict[f] + str(i+2)] = cycles
            total_cycles = sum([int(str(sh[a+str(i+2)].value)) for a in ['B', 'C', 'D', 'E', 'F', 'G'] if sh[a+str(i+2)].value is not None])
            sh['H'+str(i+2)] = total_cycles*int(n_layers)
        '''
        mcfs = [1, 2, 4, 8, 16]
        gen_paths = ["traces/" + model + "/GEN_mcf" + str(i) for i in mcfs]
        for gpi, gen_path in enumerate(gen_paths):

            flist = sorted(os.listdir(gen_path))
            name = model + "_GEN_mcf" + str(mcfs[gpi])
            if name in wb.get_sheet_names():
                wb.remove_sheet(wb.get_sheet_by_name(name))

            wb.create_sheet(title=name)
            sh = wb.get_sheet_by_name(name)
            sh['A'+str(1)] = 'KV_cache_len'
            sh['B'+str(1)] = 'createQKV'
            sh['C'+str(1)] = 'QK'
            sh['D'+str(1)] = 'SV'
            sh['E'+str(1)] = 'Wo'
            sh['F'+str(1)] = 'L1'
            sh['G'+str(1)] = 'L2'
            for i, f in enumerate(flist):
                    print(model, f)
                    if "QK" not in f and "SV" not in f:
                        pass
                    elif f == "createQKV":
                        pass
                    else:
                        continue

                    os.system("echo " + f + " >> logs/" + model + "_GEN_mcf" + str(mcfs[gpi]) + ".log")
                    os.system("./build/dramsim3main configs/HBM2_8Gb_x128.ini -c 5000000 -t " + gen_path + '/' + f  + " >> logs/" + model + "_GEN_mcf" + str(mcfs[gpi]) + ".log")

                    with open('dramsim3.json') as df:
                        json_object = json.load(df)
                        cycles = json_object['0']['num_cycles']
                        if cycles == 5000000:
                            raise Exception("compute not finished!")
                        if "QK_" in f or "SV_" in f:
                            l, ii = f.split('_')
                            sh['A'+ii] = int(ii)
                            sh[layer_dict[l] + ii] = cycles
                        else:
                            sh[layer_dict[f] + '2'] = cycles
`                   # * #L * BS / min(BS*H, 16)

        wb.save('result.xlsx')
    return


if __name__ == "__main__":
    args = parser.parse_args()
    run(args.p, args.n)
