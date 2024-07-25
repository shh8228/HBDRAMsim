import os
import argparse
import math
import openpyxl
import json

parser = argparse.ArgumentParser()
parser.add_argument("-p", default="traces", help="trace path")
parser.add_argument("-n", default="Sheet", help="sheet name")

def run(path, name):

    layer_dict = {"createQKV":'B', "QK":'C', "SV":'D', "L1":'F', "L2":'G'}
    if os.path.exists('result.xlsx'):
        wb = openpyxl.load_workbook('result.xlsx')
    else:
        wb = openpyxl.Workbook()

    fin = open("models", 'r')
    lines = fin.readlines()
    print(lines)
    for line in lines:
        model, n_heads, d_head, n_layers  = line.strip().split(' ')


        sum_path = "traces/" + model + "/SUM"
        slist = sorted(os.listdir(sum_path))
        name = model + "_SUM"
        if name in wb.get_sheet_names():
            wb.remove_sheet(wb.get_sheet_by_name(name))

        wb.create_sheet(title=name)
        sh = wb.get_sheet_by_name(name)

        for i, s in enumerate(slist):
            sh['A'+str(i+2)] = s
            flist = sorted(os.listdir(sum_path + '/' + s))
            for j, f in enumerate(flist):
                print(model, s, f)
                os.system("echo " + f + " >> logs/" + model + "_SUM_" + s + ".log")
                os.system("./build/dramsim3main configs/HBM2_8Gb_x128.ini -c 10000000 -t " + sum_path + '/' + s + '/' + f  + " >> logs/" + model + "_SUM_" + s + ".log")

                with open('dramsim3.json') as df:
                    json_object = json.load(df)
                    cycles = json_object['0']['num_cycles']
                    if cycles == 5000000:
                        raise Exception("compute not finished!")
                    cycles *= int(n_layers)
                    if f == "createQKV":
                        sh['E' + str(i+2)] = cycles
                        cycles *= 3
                    elif f == "QK" or f == "SV":
                        cycles *= int(n_heads)
                    sh[layer_dict[f] + str(i+2)] = cycles


        gen_paths = ["traces/" + model + "/GEN_mcf" + str(i) for i in [2, 4, 8]]
        for gpi, gen_path in enumerate(gen_paths):
            flist = sorted(os.listdir(gen_path))
            name = model + "_GEN_mcf" + str(2**gpi)
            if name in wb.get_sheet_names():
                wb.remove_sheet(wb.get_sheet_by_name(name))

            wb.create_sheet(title=name)
            sh = wb.get_sheet_by_name(name)
            for i, f in enumerate(flist):
                    print(model, f)
                    os.system("echo " + f + " >> logs/" + model + "_GEN_mcf" + str(2**gpi) + ".log")
                    os.system("./build/dramsim3main configs/HBM2_8Gb_x128.ini -c 10000000 -t " + gen_path + '/' + f  + " >> logs/" + model + "_GEN_mcf" + str(2**gpi) + ".log")

                    with open('dramsim3.json') as df:
                        json_object = json.load(df)
                        cycles = json_object['0']['num_cycles']
                        if cycles == 5000000:
                            raise Exception("compute not finished!")
                        cycles *= int(n_layers)

                        if f == "createQKV":
                            sh['E' + str(1)] = cycles
                            sh[layer_dict[f] + str(1)] = cycles*3
                        elif "QK" in f or "SV" in f:
                            l, ii = f.split('_')
                            sh['A'+ii] = int(ii)
                            sh[layer_dict[l] + ii] = cycles*int(n_heads)
                        else:
                            sh[layer_dict[f] + str(1)] = cycles


        wb.save('result.xlsx')
    return


if __name__ == "__main__":
    args = parser.parse_args()
    run(args.p, args.n)
